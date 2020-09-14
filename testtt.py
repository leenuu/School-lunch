from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from PIL import Image
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5 import uic
import datetime as dy
import requests
import urllib.request
import cv2
import os 
import sys

class food_time(object):

        
    def setupUi(self, Form):
        ft = "굴림"
        sz = 16
        self.lunch = self.food()
        self.td = self.time_tb()
        
        Form.setObjectName("Form")
        Form.setFixedSize(640, 700)
        Form.resize(640, 700)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(Form.sizePolicy().hasHeightForWidth())
        Form.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(False)
        font.setWeight(50)
        Form.setFont(font)
        self.gridLayout = QtWidgets.QGridLayout(Form)
        self.gridLayout.setObjectName("gridLayout")
        self.tabWidget = QtWidgets.QTabWidget(Form)
        self.tabWidget.setLayoutDirection(QtCore.Qt.LeftToRight)
        self.tabWidget.setObjectName("tabWidget")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.img_label = QtWidgets.QLabel(self.tab)
        self.img_label.setGeometry(QtCore.QRect(1, 323, 611, 320))
        self.img_label.setAlignment(QtCore.Qt.AlignCenter)
        self.img_label.setObjectName("img_label")
        self.food_label = QtWidgets.QLabel(self.tab)
        self.food_label.setGeometry(QtCore.QRect(0, 0, 611, 320))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.food_label.sizePolicy().hasHeightForWidth())
        self.food_label.setSizePolicy(sizePolicy)
        self.food_label.setAlignment(QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.food_label.setObjectName("food_label")
        self.refr = QtWidgets.QPushButton(self.tab)
        self.refr.setGeometry(QtCore.QRect(570, 10, 31, 31))
        self.refr.setObjectName("refr")
        self.tabWidget.addTab(self.tab, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.time_label = QtWidgets.QLabel(self.tab_2)
        self.time_label.setGeometry(QtCore.QRect(0, 190, 611, 511))
        self.time_label.setText("")
        self.time_label.setAlignment(QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.time_label.setObjectName("time_label")
        self.comboBox = QtWidgets.QComboBox(self.tab_2)
        self.comboBox.setGeometry(QtCore.QRect(150, 90, 111, 41))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.comboBox.sizePolicy().hasHeightForWidth())
        self.comboBox.setSizePolicy(sizePolicy)
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.pushButton = QtWidgets.QPushButton(self.tab_2)
        self.pushButton.setGeometry(QtCore.QRect(340, 90, 111, 41))
        self.pushButton.setObjectName("pushButton")
        self.text = QtWidgets.QLabel(self.tab_2)
        self.text.setGeometry(QtCore.QRect(150, -10, 301, 91))
        self.text.setAlignment(QtCore.Qt.AlignCenter)
        self.text.setObjectName("text")
        self.tabWidget.addTab(self.tab_2, "")
        self.gridLayout.addWidget(self.tabWidget, 0, 0, 1, 1)
        myfont = QtGui.QFont(ft, sz)
        myfont.setBold(True)
        self.refr.setFont(QtGui.QFont(ft, 13))
        self.refr.setToolTip('음식 사진 불러오기(자주 누르지 마세요.)')
        self.food_label.setFont(myfont)
        self.time_label.setFont(myfont)
        self.img_label.setFont(myfont)
        self.text.setFont(myfont)
        self.retranslateUi(Form)
        self.tabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(Form)

        self.pushButton.clicked.connect(self.get_time)
        self.refr.clicked.connect(self.reimg)
        self.food_label.setText(self.lunch[0])
        if self.lunch[1] == 0:
          self.qPixmapFileVar = QtGui.QPixmap()
          self.qPixmapFileVar.load("img.jpg")
          # self.qPixmapFileVar = self.qPixmapFileVar.scaledToWidth(600)
          self.img_label.setPixmap(self.qPixmapFileVar)
        else:
          self.img_label.setText("\n\n\n음식사진 업로드 안됨 \n 3교시이후 업로드 예정")
    
    def reimg(self):
      img_er = 0
      try:
        url = f'http://jeil.jje.hs.kr/jeil-h/food/2020/{dy.datetime.today().month}/{dy.datetime.today().day}/lunch'
        ua = UserAgent()
        header = {'User-Agent':str(ua.chrome)}
        req_html = requests.get(url, headers=header)
        html = req_html.text

        soup = BeautifulSoup(html,'html.parser')
        img_ = 'http://jeil.jje.hs.kr/'+ soup.find('img').get('src')

        urllib.request.urlretrieve(img_,'img.jpg')

        image = Image.open('img.jpg')

        resize_image = image.resize((611,345))

        resize_image.save('img.jpg')
      
      except AttributeError:
        img_er = 1

      if img_er == 0:
        self.qPixmapFileVar = QtGui.QPixmap()
        self.qPixmapFileVar.load("img.jpg")
        # self.qPixmapFileVar = self.qPixmapFileVar.scaledToWidth(600)
        self.img_label.setPixmap(self.qPixmapFileVar)

    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "급식 및 시간표"))
        self.img_label.setText(_translate("Form", "TextLabel"))
        self.food_label.setText(_translate("Form", "TextLabel"))
        self.refr.setText(_translate("Form", "🔄"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab), _translate("Form", "    급식    "))
        self.comboBox.setItemText(0, _translate("Form", "1반"))
        self.comboBox.setItemText(1, _translate("Form", "2반"))
        self.comboBox.setItemText(2, _translate("Form", "3반"))
        self.comboBox.setItemText(3, _translate("Form", "4반"))
        self.comboBox.setItemText(4, _translate("Form", "5반"))
        self.comboBox.setItemText(5, _translate("Form", "6반"))
        self.comboBox.setItemText(6, _translate("Form", "7반"))
        self.comboBox.setItemText(7, _translate("Form", "8반"))
        self.comboBox.setItemText(8, _translate("Form", "9반"))
        self.comboBox.setItemText(9, _translate("Form", "10반"))
        self.comboBox.setItemText(10, _translate("Form", "11반"))
        self.comboBox.setItemText(11, _translate("Form", "12반"))
        self.comboBox.setItemText(12, _translate("Form", "13반"))
        self.pushButton.setText(_translate("Form", "조회"))
        self.text.setText(_translate("Form", "시간표"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.tab_2), _translate("Form", "    시간표    "))

    def get_time(self):
      try:
          print(self.comboBox.currentText())
          print(self.td[1])
          print(f"시간표 {self.td[1]}.xlsx")
          data = load_workbook(f"시간표 {self.td[1]}.xlsx", data_only=True)
          print(data)
          ws = data.active

          time = ''

          for p in range(1,14):
            #   print(1)
              if self.comboBox.currentText() == f'{p}반':
                #   print(2)
                  for q in range(2,9-self.st):
                    # print(3)
                    if str(ws.cell(row=q, column=p+2).value).count('/') >= 4:
                      sttt = 0
                      strr = ''
                      for i in str(ws.cell(row=q, column=p+2).value).split('/'):
                        sttt += 1
                        if i == str(ws.cell(row=q, column=p+2).value).split('/')[len(str(ws.cell(row=q, column=p+2).value).split('/'))-1]:
                          strr += i 
                        else:
                          strr += i + '/'
                        if sttt == 5:
                          strr += '\n'
                          sttt = 0
                      time += f'{q-1}교시:' + strr
                    else:
                      time += f'{q-1}교시:' + str(ws.cell(row=q, column=p+2).value).replace('\n', ' ') + '\n'
                    
                      
          #   for q in range(3,10):
          #         time += f'{q-2}교시:' + str(ws.cell(row=q, column=15).value).replace('\n',' ') + '\n'

          self.time_label.setText(time)

      except NameError:
        self.time_label.setText("시간표 업로드 안됨")

    
    def food(self):
        # data = list()
        text_data = '\n'
        img_er = 0
        text = ''
        # text_ = ''
        te = ''
        url = f'http://jeil.jje.hs.kr/jeil-h/food/2020/{dy.datetime.today().month}/{dy.datetime.today().day}/lunch'

        ua = UserAgent()
        header = {'User-Agent':str(ua.chrome)}
        req_html = requests.get(url, headers=header)
        html = req_html.text

        soup = BeautifulSoup(html,'html.parser')

        try:
            img_ = 'http://jeil.jje.hs.kr/'+ soup.find('img').get('src')
            urllib.request.urlretrieve(img_,'img.jpg')
            image = Image.open('img.jpg')
            resize_image = image.resize((611,345))
            resize_image.save('img.jpg')

        except AttributeError:
            img_er = 1

        te = str(soup.find_all('dd')[1]).replace('<br/>','\n').replace('<dd>','').replace('</dd>','')

        # print(te)
        for i in te:
            try:
                int(i)
            except ValueError:
                if i != '/' :
                    text += i 
        
        text = text.split('\n')
        # print(text)

        for i in range(len(text)):
            if text[i] != '' :
                if ')' in text[i]:
                    text_data += text[i][:text[i].index(")")+1] + '\n'
                else:
                    text_data += text[i] + '\n'
            else:
                text_data += '\n' 

        if '%' in text_data:
        #   print('1')
            text_data = text_data.replace('%','')

        if '.' in text_data:
        #   print('2')
            text_data = text_data.replace('.','')

        text_data = f'{dy.datetime.today().month}월 {dy.datetime.today().day}일 점심 식단\n' + text_data 

        return [text_data, img_er]

    def time_tb(self):
        sd_data = ''
        get_sd_url = f'http://jeil.jje.hs.kr/jeil-h/0208/board/16996/'
        time_er = 0
        try:
            ua = UserAgent()
            header = {'User-Agent':str(ua.chrome)}
            req_html = requests.get(get_sd_url, headers=header)
            html = req_html.text

            soup = BeautifulSoup(html,'html.parser')

            for i in soup.find('tbody').find_all('a'):
                if f'{dy.datetime.today().month}/{dy.datetime.today().day}' in str(i): 
                # if f'{7}/{31}' in str(i): 
                    sd_data = i

            data_url = str(sd_data.get('onclick'))[str(sd_data.get('onclick')).index('(')+1:str(sd_data.get('onclick')).index(')')].split(',')[1].replace("'",'')


            url_sd = f'http://jeil.jje.hs.kr/jeil-h/0208/board/16996/{data_url}'
            print(url_sd)


            req_html = requests.get(url_sd, headers=header)
            html = req_html.text

            soup = BeautifulSoup(html,'html.parser')
            pr = soup.find('dd').find_all('a')[1].get('href')

            sh_name_day = soup.find('dd').find_all('a')[0].get_text()[soup.find('dd').find_all('a')[0].get_text().index("(") : soup.find('dd').find_all('a')[0].get_text().index(")")+1]
            sh_name = f'{dy.datetime.today().month}.{dy.datetime.today().day}{sh_name_day}'
            # sh_name = '날짜'
            # print(sh_name)

            preview_url = f'http://jeil.jje.hs.kr{pr}'

            urllib.request.urlretrieve(preview_url, 'sd.xlsx')

            data = load_workbook("sd.xlsx", data_only=True)
            ds = data.active

            wb = Workbook()
            ws = wb.active

            self.st = 0

            if ds.cell(row=10, column=2).value == None:
                self.st = 1

            for i in range(2,4):
                for j in range(3,11 - self.st):
                    ws.cell(row=j-2, column=i-1).value = ds.cell(row=j, column=i).value 
                    ws.cell(row=j-2, column=i-1).font = Font(name='맑은 고딕', size=16, bold=True)
                    ws.cell(row=j-2, column=i-1).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
                    ws.column_dimensions['B'].width = 20

            for i in range(16,29):
                for j in range(3,11 - self.st):
                    if j != 10:
                        ws.row_dimensions[j].height = 74
                    if ds.cell(row=j, column=i).value == None:
                        ws.cell(row=j-2, column=i-13).value = str(ws.cell(row=j-2, column=i-14).value)[0:2] + str(ws.cell(row=j-2, column=i-14).value)[2:]
                    else:
                        ws.cell(row=j-2, column=i-13).value = str(ds.cell(row=j, column=i).value)[0:2] + str(ds.cell(row=j, column=i).value)[2:]

                    ws.cell(row=j-2, column=i-13).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
                    ws.cell(row=j-2, column=i-13).font = Font(name='맑은 고딕', size=14, bold=True)

                wb.save(f"시간표 {sh_name}.xlsx")

        except AttributeError:
            sh_name = ''
            print("시간표 업데이트 안됨")
            time_er = 1

        return [time_er, sh_name]

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Form = QtWidgets.QWidget()
    ui = food_time()
    ui.setupUi(Form)
    Form.show()
    app.exec_()
    if ui.lunch[1] == 0:
      os.remove('img.jpg')
    if ui.td[0] == 0:
      os.remove('sd.xlsx')
      os.remove(f"시간표 {ui.td[1]}.xlsx")