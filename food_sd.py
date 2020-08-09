from fake_useragent import UserAgent
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from PIL import Image
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5 import uic
import datetime as dy
import requests
import urllib.request
import cv2
import os 
import sys

#################################################### food ######################################################################################

data = list()
text_data = '\n'
text = ''
text_ = ''
te = ''
er = 0
url = f'http://jeil.jje.hs.kr/jeil-h/food/2020/{dy.datetime.today().month}/{dy.datetime.today().day}/lunch'
# url = "http://jeil.jje.hs.kr/jeil-h/food/2020/7/31/lunch"

ua = UserAgent()
header = {'User-Agent':str(ua.chrome)}
req_html = requests.get(url, headers=header)
html = req_html.text

soup = BeautifulSoup(html,'html.parser')
try:
  img_ = 'http://jeil.jje.hs.kr/'+ soup.find('img').get('src')

  urllib.request.urlretrieve(img_,'img.jpg')

  image = Image.open('img.jpg')

  resize_image = image.resize((360,360))

  resize_image.save('img.jpg')
  
except AttributeError:
  imgg = ''
  er = 1

te = str(soup.find_all('dd')[1]).replace('<br/>','\n').replace('<dd>','').replace('</dd>','')

# print(te)
for i in te:
  try:
    int(i)
  except ValueError:
    if i != '/' :
      text += i 
  
text = text.split('\n')
# print(text)

for i in range(len(text)):
  if text[i] != '' :
    if ')' in text[i]:
      text_data += text[i][:text[i].index(")")+1] + '\n'
    else:
      text_data += text[i] + '\n'
  else:
    text_data += '\n' 

if '%' in text_data:
#   print('1')
  text_data = text_data.replace('%','')

if '.' in text_data:
#   print('2')
  text_data = text_data.replace('.','')

text_data = f'{dy.datetime.today().month}월 {dy.datetime.today().day}일 점심 식단\n' + text_data


#################################################### time_table ######################################################################################

sd_data = ''
get_sd_url = f'http://jeil.jje.hs.kr/jeil-h/0208/board/16996/'

try:
    ua = UserAgent()
    header = {'User-Agent':str(ua.chrome)}
    req_html = requests.get(get_sd_url, headers=header)
    html = req_html.text

    soup = BeautifulSoup(html,'html.parser')

    for i in soup.find('tbody').find_all('a'):
        if f'{dy.datetime.today().month}/{dy.datetime.today().day}' in str(i): 
        # if f'{7}/{31}' in str(i): 
            sd_data = i

    data_url = str(sd_data.get('onclick'))[str(sd_data.get('onclick')).index('(')+1:str(sd_data.get('onclick')).index(')')].split(',')[1].replace("'",'')


    url_sd = f'http://jeil.jje.hs.kr/jeil-h/0208/board/16996/{data_url}'
    print(url_sd)


    req_html = requests.get(url_sd, headers=header)
    html = req_html.text

    soup = BeautifulSoup(html,'html.parser')
    pr = soup.find('dd').find_all('a')[1].get('href')

    sh_name_day = soup.find('dd').find_all('a')[0].get_text()[soup.find('dd').find_all('a')[0].get_text().index("(") : soup.find('dd').find_all('a')[0].get_text().index(")")+1]
    sh_name = f'{dy.datetime.today().month}.{dy.datetime.today().day}{sh_name_day}'
    # sh_name = '날짜'
    # print(sh_name)

    preview_url = f'http://jeil.jje.hs.kr{pr}'

    urllib.request.urlretrieve(preview_url, 'sd.xlsx')

    data = load_workbook("sd.xlsx", data_only=True)
    ds = data.active

    wb = Workbook()
    ws = wb.active

    sum_cell = list()


    for i in range(2,4):
        for j in range(3,11):
            sst = ''
            ws.cell(row=j-1, column=i-1).value = ds.cell(row=j, column=i).value 
            ws.cell(row=j-1, column=i-1).font = Font(name='맑은 고딕', size=16, bold=True)
            ws.cell(row=j-1, column=i-1).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
            ws.column_dimensions['B'].width = 20

    for i in range(16,29):
        for j in range(3,11):
            if j != 10:
                ws.row_dimensions[j].height = 74
            ws.cell(row=j-1, column=i-13).value = str(ds.cell(row=j, column=i).value)[0:2] + str(ds.cell(row=j, column=i).value)[2:]
            ws.cell(row=j-1, column=i-13).alignment = Alignment(horizontal='center', vertical='center',wrapText=True)
            ws.cell(row=j-1, column=i-13).font = Font(name='맑은 고딕', size=14, bold=True)
            if ds.cell(row=j, column=i).value == None:
                sum_cell.append([i-13, j])

    # print(sum_cell)

    if sum_cell != []:
        for i in range(sum_cell[0][1]-1, sum_cell[len(sum_cell)-1][1]):
            for j in range(sum_cell[0][0], sum_cell[len(sum_cell)-1][0]+1):
                ws.cell(row=i, column=j).value = ws.cell(row=sum_cell[0][1]-1, column=sum_cell[0][0]-1).value
        
    #   ws.merge_cells(start_row= sum_cell[0][1]-1, start_column=sum_cell[0][0]-1,end_row= sum_cell[len(sum_cell)-1][1]-1,end_column=sum_cell[len(sum_cell)-1][0])
    wb.save(f"시간표 {sh_name}.xlsx")

except AttributeError:
    print("시간표 업데이트 안됨")

#################################################### gui ######################################################################################
ft = "굴림"
sz = 14

class Ui_Form(object):
    def setupUi(self, Form):
        Form.setObjectName("Form")
        Form.resize(878, 360)
        self.classnum = QtWidgets.QComboBox(Form)
        self.classnum.setGeometry(QtCore.QRect(10, 40, 71, 20))
        self.classnum.setObjectName("classnum")
        self.classnum.addItem("")
        self.classnum.addItem("")
        self.classnum.addItem("")
        self.classnum.addItem("")
        self.classnum.addItem("")
        self.classnum.addItem("")
        self.classnum.addItem("")
        self.classnum.addItem("")
        self.classnum.addItem("")
        self.classnum.addItem("")
        self.classnum.addItem("")
        self.classnum.addItem("")
        self.classnum.addItem("")
        self.time_label = QtWidgets.QLabel(Form)
        self.time_label.setGeometry(QtCore.QRect(10, 80, 201, 211))
        self.time_label.setText("")
        self.time_label.setAlignment(QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.time_label.setObjectName("time_label")
        self.pushButton = QtWidgets.QPushButton(Form)
        self.pushButton.setGeometry(QtCore.QRect(130, 40, 81, 23))
        self.pushButton.setObjectName("pushButton")
        self.food_label = QtWidgets.QLabel(Form)
        self.food_label.setGeometry(QtCore.QRect(250, 80, 231, 211))
        self.food_label.setAlignment(QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.food_label.setObjectName("food_label")
        self.img_label = QtWidgets.QLabel(Form)
        self.img_label.setGeometry(QtCore.QRect(510, 0, 360, 360))
        self.img_label.setAlignment(QtCore.Qt.AlignHCenter|QtCore.Qt.AlignTop)
        self.img_label.setObjectName("img_label")
        self.retranslateUi(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)
        self.pushButton.clicked.connect(self.get_time)
        self.food_label.setFont(QtGui.QFont(ft, sz))
        self.time_label.setFont(QtGui.QFont(ft, sz))
        self.img_label.setFont(QtGui.QFont(ft, sz))
        self.food_label.setText(text_data)
        if er == 0:
          self.qPixmapFileVar = QtGui.QPixmap()
          self.qPixmapFileVar.load("img.jpg")
          # self.qPixmapFileVar = self.qPixmapFileVar.scaledToWidth(600)
          self.img_label.setPixmap(self.qPixmapFileVar)
        else:
          self.img_label.setText("\n\n\n음식사진 업로드 안됨 \n 3교시이후 업로드 예정")


    def get_time(self):
      try:
          print(self.classnum.currentText())
          data = load_workbook(f"시간표 {sh_name}.xlsx", data_only=True)
          ws = data.active

          time = ''

          for p in range(1,14):
              if self.classnum.currentText() == f'{p}반':
                  for q in range(3,10):
                      if 15 <= len(str(ws.cell(row=q, column=p+2).value)):
                          time += f'{q-2}교시:' + str(ws.cell(row=q, column=p+2).value)[:9] 
                      else:
                          time += f'{q-2}교시:' + str(ws.cell(row=q, column=p+2).value).replace('\n',' ')[:3] + '\n'

          #   for q in range(3,10):
          #         time += f'{q-2}교시:' + str(ws.cell(row=q, column=15).value).replace('\n',' ') + '\n'

          self.time_label.setText(time)

      except NameError:
        self.time_label.setText("시간표 업로드 안됨")



    def retranslateUi(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.classnum.setItemText(0, _translate("Form", "1반"))
        self.classnum.setItemText(1, _translate("Form", "2반"))
        self.classnum.setItemText(2, _translate("Form", "3반"))
        self.classnum.setItemText(3, _translate("Form", "4반"))
        self.classnum.setItemText(4, _translate("Form", "5반"))
        self.classnum.setItemText(5, _translate("Form", "6반"))
        self.classnum.setItemText(6, _translate("Form", "7반"))
        self.classnum.setItemText(7, _translate("Form", "8반"))
        self.classnum.setItemText(8, _translate("Form", "9반"))
        self.classnum.setItemText(9, _translate("Form", "10반"))
        self.classnum.setItemText(10, _translate("Form", "11반"))
        self.classnum.setItemText(11, _translate("Form", "12반"))
        self.classnum.setItemText(12, _translate("Form", "13반"))
        self.pushButton.setText(_translate("Form", "조회"))
        self.food_label.setText(_translate("Form", "TextLabel"))
        self.img_label.setText(_translate("Form", "img"))



if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    Form = QtWidgets.QWidget()
    ui = Ui_Form()
    ui.setupUi(Form)
    Form.show()
    sys.exit(app.exec_())

    os.remove('img.jpg')
    os.remove('sd.xlsx')
    os.remove(f"시간표 {sh_name}.xlsx")



